import logging
from datetime import date, datetime

from openpyxl import load_workbook

log = logging.getLogger(__name__)

LOCATION_SHEETS   = ['Blanchardstown', 'Cork', 'Liffey Valley', 'Nutgrove', 'Whitewater']
ACTIVE_SHEETS     = ['Blanchardstown', 'Cork', 'Nutgrove']
INCOME_ROW        = 5
DEFAULT_RATE_WDAY = 14.15
DEFAULT_RATE_SUN  = 17.98

COL_NAME         = 3   # C
COL_RATE_MON     = 4   # D
COL_RATE_SUN_COL = 5   # E
COL_TOT_WDAY     = 18  # R
COL_TOT_WDAY_PAY = 19  # S
COL_TOT_SUN      = 20  # T
COL_TOT_SUN_PAY  = 21  # U
COL_BONUS        = 22  # V
COL_ADJ          = 23  # W
COL_FINAL        = 24  # X

DATA_COL_START = 6   # F
DATA_COL_END   = 16  # P
PERIOD_ROW     = 2


def _norm(name: str) -> str:
    return name.strip().lower()


def _num(v):
    try:
        return float(v) if v is not None and not isinstance(v, str) else 0.0
    except (TypeError, ValueError):
        return 0.0


def _sheet_col_types(ws) -> dict:
    """
    Return {col_index: 'weekday'|'cross_month'|'sunday'|None} by reading row 2 labels.
    'Sunday' columns have label 'Sun'.
    All other labelled columns are weekday (including cross-month splits).
    """
    result = {}
    for col in range(DATA_COL_START, DATA_COL_END + 1):
        label = ws.cell(PERIOD_ROW, col).value
        if label == 'Sun':
            result[col] = 'sunday'
        elif label:
            result[col] = 'weekday'
        else:
            result[col] = None
    return result


def _find_week_columns(ws, target_sunday: date):
    """
    Find column indices for a given week's Sunday.

    For a normal week: [Mon-Sat col] [Sun col]
    For a cross-month week: [weekday col] [cross_month col] [Sun col]

    Detection: if the column immediately before the Sunday is NOT 'Mon-Sat' and NOT 'Sun',
    AND the column before that is also a labelled non-Sunday column → cross-month split.

    Returns (weekday_col, sunday_col, cross_month_col) where cross_month_col may be None.
    """
    for cell in ws[3]:
        if isinstance(cell.value, datetime) and cell.value.date() == target_sunday:
            sunday_col = cell.column
            prev_label      = ws.cell(PERIOD_ROW, sunday_col - 1).value or ''
            prev_prev_label = ws.cell(PERIOD_ROW, sunday_col - 2).value or ''

            # Cross-month: two non-Sunday, non-Mon-Sat columns before the Sunday
            is_cross = (
                prev_label not in ('Mon-Sat', 'Sun', '') and
                prev_prev_label not in ('Sun', '')
            )

            if is_cross:
                cross_month_col = sunday_col - 1
                weekday_col     = sunday_col - 2
                log.info(
                    f"[{ws.title}] Sunday {target_sunday} → col {sunday_col} | "
                    f"cross-month col {cross_month_col} | weekday col {weekday_col}"
                )
                return weekday_col, sunday_col, cross_month_col
            else:
                weekday_col = sunday_col - 1
                log.info(f"[{ws.title}] Sunday {target_sunday} → col {sunday_col} | weekday col {weekday_col}")
                return weekday_col, sunday_col, None

    log.warning(f"[{ws.title}] No column found for Sunday {target_sunday}")
    return None, None, None


def _get_sheet_names(ws) -> list:
    names = []
    for row in ws.iter_rows(min_row=7, min_col=COL_NAME, max_col=COL_NAME):
        cell = row[0]
        if cell.value and isinstance(cell.value, str) and cell.value.strip():
            names.append(cell.value.strip())
    return names


def _find_last_staff_row(ws) -> int:
    last = 7
    for row in ws.iter_rows(min_row=7, max_row=50, min_col=COL_NAME, max_col=COL_NAME):
        if row[0].value and isinstance(row[0].value, str) and row[0].value.strip():
            last = row[0].row
    return last


def _find_staff_row(ws, name: str):
    target = _norm(name)
    for row in ws.iter_rows(min_row=7, min_col=COL_NAME, max_col=COL_NAME):
        cell = row[0]
        if cell.value and _norm(str(cell.value)) == target:
            return cell.row
    return None


def _match_name(square_name: str, sheet_names: list):
    sq_norm  = _norm(square_name)
    sq_parts = sq_norm.split()
    sheet_map = {_norm(n): n for n in sheet_names}
    if sq_norm in sheet_map:
        return sheet_map[sq_norm]
    if sq_parts and sq_parts[0] in sheet_map:
        return sheet_map[sq_parts[0]]
    if len(sq_parts) > 1 and sq_parts[-1] in sheet_map:
        return sheet_map[sq_parts[-1]]
    return None


def _add_new_staff_row(ws, name: str, new_row: int):
    r = new_row
    ws.cell(r, COL_NAME).value         = name
    ws.cell(r, COL_RATE_MON).value     = DEFAULT_RATE_WDAY
    ws.cell(r, COL_RATE_SUN_COL).value = DEFAULT_RATE_SUN

    # Build formulas based on actual column layout (handles split columns)
    col_types = _sheet_col_types(ws)
    wday_cols = [col for col, t in col_types.items() if t == 'weekday']
    sun_cols  = [col for col, t in col_types.items() if t == 'sunday']

    def col_letter(c):
        return chr(64 + c) if c <= 26 else chr(64 + c // 26) + chr(64 + c % 26)

    wday_formula = '+'.join(f'{col_letter(c)}{r}' for c in wday_cols) or '0'
    sun_formula  = '+'.join(f'{col_letter(c)}{r}' for c in sun_cols)  or '0'

    ws.cell(r, COL_TOT_WDAY).value     = f'={wday_formula}'
    ws.cell(r, COL_TOT_WDAY_PAY).value = f'=R{r}*$D{r}'
    ws.cell(r, COL_TOT_SUN).value      = f'={sun_formula}'
    ws.cell(r, COL_TOT_SUN_PAY).value  = f'=T{r}*$E{r}'
    ws.cell(r, COL_FINAL).value        = f'=S{r}+U{r}+V{r}+W{r}'
    log.info(f"[{ws.title}] Added '{name}' at row {r} with default rates")


def _write_staff_totals(ws, row: int, rate_mon: float, rate_sun: float):
    """
    Compute and hard-write summary columns (R, S, T, U, X).
    Reads column types dynamically so split columns are handled correctly.
    ALL non-Sunday labelled columns count as weekday hours (including cross-month).
    """
    col_types  = _sheet_col_types(ws)
    total_wday = sum(_num(ws.cell(row, c).value) for c, t in col_types.items() if t == 'weekday')
    total_sun  = sum(_num(ws.cell(row, c).value) for c, t in col_types.items() if t == 'sunday')
    bonus      = _num(ws.cell(row, COL_BONUS).value)
    adj        = _num(ws.cell(row, COL_ADJ).value)

    wday_pay = total_wday * rate_mon
    sun_pay  = total_sun  * rate_sun
    final    = wday_pay + sun_pay + bonus + adj

    ws.cell(row, COL_TOT_WDAY).value     = round(total_wday, 2)
    ws.cell(row, COL_TOT_WDAY_PAY).value = round(wday_pay, 2)
    ws.cell(row, COL_TOT_SUN).value      = round(total_sun, 2)
    ws.cell(row, COL_TOT_SUN_PAY).value  = round(sun_pay, 2)
    ws.cell(row, COL_FINAL).value        = round(final, 2)


def _clear_inactive_sheet(ws):
    for row in ws.iter_rows(min_row=7, max_row=50):
        name = ws.cell(row[0].row, COL_NAME).value
        if not name or not isinstance(name, str):
            continue
        r = row[0].row
        for col in range(DATA_COL_START, DATA_COL_END + 1):
            ws.cell(r, col).value = None
        for col in [COL_TOT_WDAY, COL_TOT_WDAY_PAY, COL_TOT_SUN, COL_TOT_SUN_PAY, COL_FINAL]:
            ws.cell(r, col).value = 0


def _clear_week_columns(ws, weekday_col: int, sunday_col: int, cross_month_col=None):
    cols_to_clear = {weekday_col, sunday_col}
    if cross_month_col:
        cols_to_clear.add(cross_month_col)
    for col in cols_to_clear:
        for row in ws.iter_rows(min_row=INCOME_ROW, min_col=col, max_col=col):
            cell = row[0]
            if not isinstance(cell.value, str):
                cell.value = None


def update_excel_wages(
    file_path: str,
    shifts: dict,
    income: dict,
    target_sunday: date = None,
) -> dict:
    if target_sunday is None:
        from square_service import get_previous_week_sunday
        target_sunday = get_previous_week_sunday()

    wb      = load_workbook(file_path)
    summary = {}

    # Zero out inactive sheets
    for sheet_name in LOCATION_SHEETS:
        if sheet_name not in ACTIVE_SHEETS and sheet_name in wb.sheetnames:
            _clear_inactive_sheet(wb[sheet_name])

    for sheet_name in ACTIVE_SHEETS:
        if sheet_name not in wb.sheetnames:
            log.warning(f"Sheet '{sheet_name}' missing — skipping")
            continue

        ws              = wb[sheet_name]
        location_shifts = shifts.get(sheet_name, {})
        location_income = income.get(sheet_name, {})

        if isinstance(location_income, dict):
            income_primary = location_income.get('income', 0.0)
            income_cross   = location_income.get('cross_month_income', 0.0)
        else:
            income_primary = float(location_income or 0)
            income_cross   = 0.0

        weekday_col, sunday_col, cross_month_col = _find_week_columns(ws, target_sunday)
        if not weekday_col:
            summary[sheet_name] = {'error': f'Column not found for Sunday {target_sunday}'}
            continue

        _clear_week_columns(ws, weekday_col, sunday_col, cross_month_col)

        # Write income
        if cross_month_col:
            # Separate columns available — split the income
            if income_primary:
                ws.cell(row=INCOME_ROW, column=weekday_col).value = income_primary
            if income_cross:
                ws.cell(row=INCOME_ROW, column=cross_month_col).value = income_cross
        else:
            # No separate cross-month column — combine income into one
            total_income = income_primary + income_cross
            if total_income:
                ws.cell(row=INCOME_ROW, column=weekday_col).value = round(total_income, 2)

        if not location_shifts:
            summary[sheet_name] = {
                'updated': [], 'unmatched': [], 'added': [],
                'note': 'No shifts this week',
                'income': income_primary, 'cross_month_income': income_cross,
            }
            for name in _get_sheet_names(ws):
                row = _find_staff_row(ws, name)
                if row:
                    _write_staff_totals(ws, row,
                                        _num(ws.cell(row, COL_RATE_MON).value),
                                        _num(ws.cell(row, COL_RATE_SUN_COL).value))
            continue

        sheet_names = _get_sheet_names(ws)
        updated   = []
        unmatched = []
        added     = []

        for sq_name, hours in location_shifts.items():
            matched = _match_name(sq_name, sheet_names)

            if not matched:
                new_row = _find_last_staff_row(ws) + 1
                _add_new_staff_row(ws, sq_name, new_row)
                sheet_names.append(sq_name)
                matched = sq_name
                added.append(sq_name)

            row = _find_staff_row(ws, matched)
            if not row:
                unmatched.append(sq_name)
                continue

            weekday_h     = round(hours['weekday_hours'], 2)
            sunday_h      = round(hours['sunday_hours'],  2)
            cross_month_h = round(hours.get('cross_month_hours', 0.0), 2)

            if weekday_h > 0:
                ws.cell(row=row, column=weekday_col).value = weekday_h
            if sunday_h > 0:
                ws.cell(row=row, column=sunday_col).value  = sunday_h
            if cross_month_h > 0:
                if cross_month_col:
                    ws.cell(row=row, column=cross_month_col).value = cross_month_h
                else:
                    # No separate column — add to weekday total
                    existing = _num(ws.cell(row=row, column=weekday_col).value)
                    ws.cell(row=row, column=weekday_col).value = round(existing + cross_month_h, 2)

            updated.append({
                'name':              matched,
                'weekday_hours':     weekday_h,
                'sunday_hours':      sunday_h,
                'cross_month_hours': cross_month_h,
            })
            log.info(f"[{sheet_name}] {matched}: {weekday_h}h weekday, "
                     f"{cross_month_h}h cross-month, {sunday_h}h Sun")

        # Recompute totals for ALL staff
        for name in sheet_names:
            row = _find_staff_row(ws, name)
            if row:
                _write_staff_totals(ws, row,
                                    _num(ws.cell(row, COL_RATE_MON).value),
                                    _num(ws.cell(row, COL_RATE_SUN_COL).value))

        summary[sheet_name] = {
            'updated':            updated,
            'unmatched':          unmatched,
            'added':              added,
            'income':             income_primary,
            'cross_month_income': income_cross,
        }
        log.info(f"[{sheet_name}] Done — {len(updated)} updated, {len(added)} added, "
                 f"{len(unmatched)} unmatched | income €{income_primary:.2f} + €{income_cross:.2f}")

    wb.save(file_path)
    log.info(f"Saved workbook to {file_path}")
    return summary


def write_prev_month_end(file_path: str, shifts: dict, income: dict) -> dict:
    """
    Write the in-month weekday portion of a cross-month week to the PREVIOUS month's file.
    Finds the last non-Sunday data column (the end partial) and writes there.
    Does NOT use Sunday datetime for column detection (Sunday is in the next month's file).
    """
    wb = load_workbook(file_path)
    summary = {}

    for sheet_name in ACTIVE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue

        ws              = wb[sheet_name]
        location_shifts = shifts.get(sheet_name, {})
        location_income = income.get(sheet_name, {})
        income_primary  = location_income.get('income', 0.0) if isinstance(location_income, dict) else float(location_income or 0)

        # Find the last non-Sunday data column (the partial end column)
        end_col = None
        for col in range(DATA_COL_START, DATA_COL_END + 1):
            label = ws.cell(PERIOD_ROW, col).value
            if label and label != 'Sun':
                end_col = col

        if not end_col:
            log.warning(f"[{sheet_name}] Could not find end partial column in prev month file")
            continue

        # Clear end column
        for row_idx in range(INCOME_ROW, 51):
            cell = ws.cell(row_idx, end_col)
            if not isinstance(cell.value, str):
                cell.value = None

        if income_primary:
            ws.cell(INCOME_ROW, end_col).value = income_primary

        if not location_shifts:
            summary[sheet_name] = {'updated': [], 'unmatched': [], 'added': [], 'note': 'No shifts', 'income': income_primary}
            for name in _get_sheet_names(ws):
                row = _find_staff_row(ws, name)
                if row:
                    _write_staff_totals(ws, row,
                                        _num(ws.cell(row, COL_RATE_MON).value),
                                        _num(ws.cell(row, COL_RATE_SUN_COL).value))
            continue

        sheet_names = _get_sheet_names(ws)
        updated   = []
        unmatched = []

        for sq_name, hours in location_shifts.items():
            weekday_h = round(hours['weekday_hours'], 2)
            if weekday_h == 0:
                continue

            matched = _match_name(sq_name, sheet_names)
            if not matched:
                unmatched.append(sq_name)
                log.warning(f"[{sheet_name}] Could not match '{sq_name}' in prev month file")
                continue

            row = _find_staff_row(ws, matched)
            if not row:
                unmatched.append(sq_name)
                continue

            ws.cell(row, end_col).value = weekday_h
            updated.append({'name': matched, 'weekday_hours': weekday_h, 'sunday_hours': 0.0})
            log.info(f"[{sheet_name}] (prev month) {matched}: {weekday_h}h weekday")

        # Recompute totals for all staff
        for name in _get_sheet_names(ws):
            row = _find_staff_row(ws, name)
            if row:
                _write_staff_totals(ws, row,
                                    _num(ws.cell(row, COL_RATE_MON).value),
                                    _num(ws.cell(row, COL_RATE_SUN_COL).value))

        summary[sheet_name] = {'updated': updated, 'unmatched': unmatched, 'added': [], 'income': income_primary}
        log.info(f"[{sheet_name}] (prev month) Done — {len(updated)} updated, {len(unmatched)} unmatched")

    wb.save(file_path)
    log.info(f"Saved prev month file: {file_path}")
    return summary
