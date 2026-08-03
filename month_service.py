"""
month_service.py — auto-generates monthly wages files from the master template.

Cross-month splits: when a Mon-Sat period crosses a month boundary the period
is emitted as TWO columns (e.g. 'Mon-Fri 27-31 Jul' + 'Sat 01 Aug') so every
day is visible and income can be attributed accurately. Splits are only applied
when the total number of columns stays within the 11-column limit (F-P). Months
with 4+ full weeks that also have cross-month boundaries on both ends cannot
fit splits and fall back to combined labels (e.g. 'Mon-Sat 27 Jul-01 Aug').
"""

import calendar
import logging
import shutil
from datetime import date, timedelta, datetime
from pathlib import Path

from openpyxl import load_workbook

log = logging.getLogger(__name__)

LOCATION_SHEETS = ['Blanchardstown', 'Cork', 'Liffey Valley', 'Nutgrove', 'Whitewater']
MONTH_ROW       = 1
MONTH_COL       = 6
PERIOD_ROW      = 2
DATE_ROW        = 3
STAFF_ROW_START = 7
STAFF_ROW_END   = 25
DATA_COL_START  = 6    # F
DATA_COL_END    = 16   # P  — max 11 data columns

MONTH_NAMES = ['', 'January', 'February', 'March', 'April', 'May', 'June',
               'July', 'August', 'September', 'October', 'November', 'December']
MONTH_ABBR  = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
DAY_ABBR    = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
MAX_COLS    = DATA_COL_END - DATA_COL_START + 1   # 11


# ── Helpers ───────────────────────────────────────────────────────────────────

def _range_label(start: date, end: date) -> str:
    if start.month == end.month:
        return f"{start.day:02d}-{end.day:02d} {MONTH_ABBR[start.month]}"
    return f"{start.day:02d} {MONTH_ABBR[start.month]}-{end.day:02d} {MONTH_ABBR[end.month]}"


def _dow_label(start: date, end: date) -> str:
    s = DAY_ABBR[start.weekday()]
    e = DAY_ABBR[end.weekday()]
    return f"{s}-{e}" if s != e else s


def _crosses_month(start: date, end: date) -> bool:
    return start.month != end.month


def _split_period(start: date, end: date):
    """If start/end cross a month boundary return (end_of_first_month, start_of_next_month)."""
    if _crosses_month(start, end):
        last = date(start.year, start.month,
                    calendar.monthrange(start.year, start.month)[1])
        return last, last + timedelta(days=1)
    return None, None


def _plan_splits(year: int, month: int):
    """
    Pre-compute whether start and/or end partial columns should be split.
    Returns (p_start, p_end, sunday, do_start_split, do_end_split).
    """
    first_day = date(year, month, 1)
    last_day  = date(year, month, calendar.monthrange(year, month)[1])
    dow       = first_day.weekday()

    if dow == 0:
        p_start = first_day - timedelta(days=3)
        p_end   = first_day - timedelta(days=2)
        sunday  = first_day - timedelta(days=1)
    else:
        p_start = first_day - timedelta(days=dow)
        p_end   = p_start + timedelta(days=5)
        sunday  = p_start + timedelta(days=6)

    start_crosses = _crosses_month(p_start, p_end)

    # Count full weeks
    full_weeks = 0
    cm = sunday + timedelta(days=1)
    while cm <= last_day:
        if cm + timedelta(days=5) > last_day:
            break
        full_weeks += 1
        cm += timedelta(days=7)

    has_end     = cm <= last_day
    end_crosses = has_end and _crosses_month(cm, cm + timedelta(days=5))

    n_s = 2 if start_crosses else 1
    n_e = (2 if end_crosses else 1) if has_end else 0
    total_both = n_s + 1 + full_weeks * 2 + n_e

    if total_both <= MAX_COLS:
        do_start = start_crosses
        do_end   = end_crosses
    elif (1 + 1 + full_weeks * 2 + (2 if end_crosses else 1 if has_end else 0)) <= MAX_COLS:
        do_start = False
        do_end   = end_crosses
    elif (n_s + 1 + full_weeks * 2 + (1 if has_end else 0)) <= MAX_COLS:
        do_start = start_crosses
        do_end   = False
    else:
        do_start = False
        do_end   = False

    return p_start, p_end, sunday, do_start, do_end


# ── Main generator ────────────────────────────────────────────────────────────

def generate_week_columns(year: int, month: int) -> list:
    """
    Build column definitions for a given month.

    Each entry: (row2_label, row3_value, col_type)
      col_type: 'weekday' | 'cross_month' | 'sunday'
    """
    first_day = date(year, month, 1)
    last_day  = date(year, month, calendar.monthrange(year, month)[1])

    p_start, p_end, sunday, do_start_split, do_end_split = _plan_splits(year, month)

    cols = []

    # ── Partial start ─────────────────────────────────────────────────────────
    split_end, split_start = _split_period(p_start, p_end)
    if do_start_split and split_end:
        cols.append((_dow_label(p_start, split_end),
                     _range_label(p_start, split_end),
                     'weekday'))
        cols.append((_dow_label(split_start, p_end),
                     _range_label(split_start, p_end),
                     'cross_month'))
    else:
        cols.append((_dow_label(p_start, p_end),
                     _range_label(p_start, p_end),
                     'weekday'))

    cols.append(('Sun',
                 datetime(sunday.year, sunday.month, sunday.day),
                 'sunday'))

    # ── Full weeks + partial end ──────────────────────────────────────────────
    current_monday = sunday + timedelta(days=1)

    while current_monday <= last_day:
        sat = current_monday + timedelta(days=5)
        sun = current_monday + timedelta(days=6)

        if sat > last_day:
            # Partial end
            split_end2, split_start2 = _split_period(current_monday, sat)
            if do_end_split and split_end2:
                cols.append((_dow_label(current_monday, split_end2),
                             _range_label(current_monday, split_end2),
                             'weekday'))
                cols.append((_dow_label(split_start2, sat),
                             _range_label(split_start2, sat),
                             'cross_month'))
            else:
                cols.append((_dow_label(current_monday, sat),
                             _range_label(current_monday, sat),
                             'weekday'))
            break
        else:
            cols.append(('Mon-Sat',
                         _range_label(current_monday, sat),
                         'weekday'))
            cols.append(('Sun',
                         datetime(sun.year, sun.month, sun.day),
                         'sunday'))

        current_monday += timedelta(days=7)

    return cols


# ── Core refresh logic ────────────────────────────────────────────────────────

def _refresh_location_sheet(ws, year: int, month: int):
    ws.cell(row=MONTH_ROW, column=MONTH_COL).value = MONTH_NAMES[month].upper()

    cols = generate_week_columns(year, month)

    for col in range(DATA_COL_START, DATA_COL_END + 1):
        ws.cell(row=PERIOD_ROW, column=col).value = None
        ws.cell(row=DATE_ROW,   column=col).value = None

    for i, (label, val, _) in enumerate(cols):
        col = DATA_COL_START + i
        ws.cell(row=PERIOD_ROW, column=col).value = label
        ws.cell(row=DATE_ROW,   column=col).value = val

    for row in range(STAFF_ROW_START, STAFF_ROW_END + 1):
        for col in range(DATA_COL_START, DATA_COL_END + 1):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell.value, str):
                cell.value = None

    log.info(f"[{ws.title}] Refreshed for {MONTH_NAMES[month]} {year} ({len(cols)} cols)")


def refresh_for_month(master_path: str, output_path: str, year: int, month: int):
    shutil.copy2(master_path, output_path)
    wb = load_workbook(output_path)

    for sheet_name in LOCATION_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        _refresh_location_sheet(wb[sheet_name], year, month)

    if 'Final Payment' in wb.sheetnames:
        dow = date(year, month, 1).weekday()
        ps  = (date(year, month, 1) - timedelta(days=3) if dow == 0
               else date(year, month, 1) - timedelta(days=dow))
        wb['Final Payment'].cell(row=1, column=2).value = datetime(
            ps.year, ps.month, ps.day)

    wb.save(output_path)
    log.info(f"Generated {Path(output_path).name}")


def monthly_filename(year: int, month: int) -> str:
    return f"MB_Ireland_{MONTH_NAMES[month][:3]}_{year}.xlsx"


def current_expected_filename() -> str:
    today = datetime.now()
    return monthly_filename(today.year, today.month)
