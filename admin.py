"""
admin.py — Admin panel for MB Ireland Wages
"""

import os
from datetime import datetime
from pathlib import Path
from functools import wraps

from flask import Blueprint, request, jsonify, make_response, redirect
from openpyxl import load_workbook

admin_bp = Blueprint('admin', __name__)

DATA_DIR        = Path(os.getenv('DATA_DIR', '/data'))
MASTER_TEMPLATE = DATA_DIR / 'master_template.xlsx'
ADMIN_PASSWORD  = os.getenv('API_KEY', 'changeme')

# Only active locations — Liffey Valley sold, Whitewater seasonal/inactive
ACTIVE_SHEETS   = ['Blanchardstown', 'Cork', 'Nutgrove']
ALL_SHEETS      = ['Blanchardstown', 'Cork', 'Liffey Valley', 'Nutgrove', 'Whitewater']

MONTH_NAMES = ['', 'January', 'February', 'March', 'April', 'May', 'June',
               'July', 'August', 'September', 'October', 'November', 'December']
MONTH_TO_NUM = {m[:3]: i for i, m in enumerate(MONTH_NAMES) if m}

COL_NAME        = 3   # C
COL_RATE_MON    = 4   # D
COL_RATE_SUN    = 5   # E
COL_TOT_WDAY    = 18  # R
COL_TOT_SUN     = 20  # T
COL_FINAL       = 24  # X
DATA_COL_START  = 6   # F
DATA_COL_END    = 16  # P
PERIOD_ROW      = 2
DATE_ROW        = 3
INCOME_ROW      = 5
STAFF_START     = 7


# ── Auth ──────────────────────────────────────────────────────────────────────

def check_auth(req):
    return req.cookies.get('admin_token') == ADMIN_PASSWORD

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not check_auth(request):
            return redirect('/admin/login')
        return f(*args, **kwargs)
    return decorated


# ── Data helpers ──────────────────────────────────────────────────────────────

def _file_date_key(f):
    """Sort by actual year+month, not alphabetically."""
    parts = f.stem.split('_')
    month_num = MONTH_TO_NUM.get(parts[2], 0) if len(parts) > 2 else 0
    year      = int(parts[3]) if len(parts) > 3 else 0
    return (year, month_num)

def get_current_file():
    files = sorted(DATA_DIR.glob('MB_Ireland_*.xlsx'), key=_file_date_key, reverse=True)
    return files[0] if files else None


def _num(v):
    try:
        return float(v) if v is not None and not isinstance(v, str) else 0.0
    except (TypeError, ValueError):
        return 0.0


def _read_col_structure(ws):
    """
    Dynamically read data column labels from row 2.
    Returns list of (col_idx, label, col_type) where col_type is 'weekday' or 'sunday'.
    """
    cols = []
    for col in range(DATA_COL_START, DATA_COL_END + 1):
        label = ws.cell(PERIOD_ROW, col).value
        if label:
            ctype = 'sunday' if label == 'Sun' else 'weekday'
            date_val = ws.cell(DATE_ROW, col).value
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%d %b')
            else:
                date_str = str(date_val) if date_val else ''
            cols.append({'col': col, 'label': label, 'date': date_str, 'type': ctype})
    return cols


def _group_into_weeks(col_structure):
    """
    Group columns into payroll weeks.
    Each week = one or more weekday cols (Mon-Sat, or split like Mon-Fri + Sat) + optional Sunday.
    """
    weeks = []
    i = 0
    while i < len(col_structure):
        entry = col_structure[i]
        if entry['type'] == 'sunday':
            i += 1
            continue

        # Collect consecutive weekday cols (handles cross-month splits)
        wday_cols = [entry]
        i += 1
        while i < len(col_structure) and col_structure[i]['type'] == 'weekday':
            wday_cols.append(col_structure[i])
            i += 1

        sun = None
        if i < len(col_structure) and col_structure[i]['type'] == 'sunday':
            sun = col_structure[i]
            i += 1

        # Build display header
        if len(wday_cols) == 1:
            header = f"{wday_cols[0]['label']}<br>{wday_cols[0]['date']}"
        else:
            header = '<br>+<br>'.join(
                f"{w['label']}<br>{w['date']}" for w in wday_cols
            )

        weeks.append({
            'wday_cols':    [w['col'] for w in wday_cols],
            'sun_col':      sun['col'] if sun else None,
            'header':       header,
            'sun_header':   f"Sun<br>{sun['date']}" if sun else '',
        })
    return weeks


def read_sheet_data(ws):
    """Read a location sheet into structured data for the live view."""
    col_structure = _read_col_structure(ws)
    weeks         = _group_into_weeks(col_structure)

    # Income per weekday column
    income = {
        c['col']: _num(ws.cell(INCOME_ROW, c['col']).value)
        for c in col_structure if c['type'] == 'weekday'
    }

    # Staff rows
    staff = []
    for row_idx in range(STAFF_START, 50):
        name = ws.cell(row_idx, COL_NAME).value
        if not name or not isinstance(name, str) or not name.strip():
            continue

        rate_mon = _num(ws.cell(row_idx, COL_RATE_MON).value)
        rate_sun = _num(ws.cell(row_idx, COL_RATE_SUN).value)
        bonus    = _num(ws.cell(row_idx, 22).value)  # V
        adj      = _num(ws.cell(row_idx, 23).value)  # W

        week_hours = {}
        for w in weeks:
            wh = sum(_num(ws.cell(row_idx, c).value) for c in w['wday_cols'])
            sh = _num(ws.cell(row_idx, w['sun_col']).value) if w['sun_col'] else 0.0
            week_hours[tuple(w['wday_cols'])] = {'wday': wh, 'sun': sh}

        total_wday = _num(ws.cell(row_idx, COL_TOT_WDAY).value)
        total_sun  = _num(ws.cell(row_idx, COL_TOT_SUN).value)
        final      = _num(ws.cell(row_idx, COL_FINAL).value)

        staff.append({
            'name':       name.strip(),
            'rate_mon':   rate_mon,
            'rate_sun':   rate_sun,
            'week_hours': week_hours,
            'total_wday': total_wday,
            'total_sun':  total_sun,
            'bonus':      bonus,
            'adj':        adj,
            'final':      final,
        })

    return {'weeks': weeks, 'income': income, 'staff': staff}


def get_staff_from_template():
    if not MASTER_TEMPLATE.exists():
        return {}
    wb = load_workbook(MASTER_TEMPLATE)
    result = {}
    for sheet in ACTIVE_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws   = wb[sheet]
        staff = []
        for row in ws.iter_rows(min_row=STAFF_START, max_row=50,
                                min_col=COL_NAME, max_col=COL_RATE_SUN):
            name = row[0].value
            if not name or not isinstance(name, str) or not name.strip():
                continue
            staff.append({
                'name':     name.strip(),
                'rate_mon': row[1].value or '',
                'rate_sun': row[2].value or '',
            })
        result[sheet] = staff
    return result


def save_staff_to_template(location, staff_list):
    if not MASTER_TEMPLATE.exists():
        return False, 'Master template not found'
    if location not in ACTIVE_SHEETS:
        return False, f'Invalid location: {location}'
    wb = load_workbook(MASTER_TEMPLATE)
    if location not in wb.sheetnames:
        return False, f'Sheet {location} not found'
    ws = wb[location]
    for row in ws.iter_rows(min_row=STAFF_START, max_row=50,
                            min_col=COL_NAME, max_col=COL_RATE_SUN):
        for cell in row:
            if not isinstance(cell.value, str):
                cell.value = None
    for i, s in enumerate(staff_list):
        r = STAFF_START + i
        ws.cell(r, COL_NAME).value    = s['name']
        ws.cell(r, COL_RATE_MON).value = float(s['rate_mon']) if s['rate_mon'] else None
        ws.cell(r, COL_RATE_SUN).value = float(s['rate_sun']) if s['rate_sun'] else None
    wb.save(MASTER_TEMPLATE)
    return True, 'Saved'


# ── Shared styles ─────────────────────────────────────────────────────────────

STYLES = '''
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',sans-serif;background:#0f1117;color:#e2e8f0;min-height:100vh}
.topbar{background:#1a1f2e;border-bottom:1px solid #2d3748;padding:0 28px;
  display:flex;align-items:center;justify-content:space-between;height:54px;
  position:sticky;top:0;z-index:100}
.logo{font-size:11px;font-weight:700;letter-spacing:.14em;text-transform:uppercase;color:#6366f1}
.topbar h1{font-size:16px;font-weight:600;color:#f1f5f9;margin-left:14px}
.topbar-left{display:flex;align-items:center}
.nav{display:flex;gap:4px;margin-left:28px}
.nav a{padding:6px 12px;border-radius:6px;font-size:13px;color:#64748b;text-decoration:none;font-weight:500}
.nav a:hover,.nav a.active{background:#1e293b;color:#e2e8f0}
.nav a.active{color:#a5b4fc}
a.logout{font-size:13px;color:#475569;text-decoration:none}
a.logout:hover{color:#94a3b8}
.main{max-width:1200px;margin:0 auto;padding:28px 20px}
.section{background:#1a1f2e;border:1px solid #2d3748;border-radius:10px;padding:22px;margin-bottom:22px}
.section-title{font-size:14px;font-weight:600;color:#f1f5f9;margin-bottom:18px}
.cards{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:22px}
.card{background:#1a1f2e;border:1px solid #2d3748;border-radius:10px;padding:18px}
.card-label{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.1em;color:#64748b;margin-bottom:10px}
.card-action{font-size:13px;color:#94a3b8;margin-bottom:14px;line-height:1.5}
.btn{display:inline-block;padding:8px 16px;border-radius:7px;font-size:13px;font-weight:500;cursor:pointer;border:none;text-decoration:none}
.btn-primary{background:#6366f1;color:#fff}.btn-primary:hover{background:#4f46e5}
.btn-secondary{background:#1e293b;border:1px solid #334155;color:#94a3b8}.btn-secondary:hover{background:#263244;color:#e2e8f0}
.btn-danger{background:#2d1f1f;border:1px solid #7f1d1d;color:#fca5a5}.btn-danger:hover{background:#3d2020}
.tabs{display:flex;gap:4px;margin-bottom:18px;flex-wrap:wrap}
.tab{padding:6px 14px;border-radius:6px;font-size:13px;font-weight:500;cursor:pointer;
  border:1px solid #2d3748;background:transparent;color:#64748b}
.tab:hover{color:#94a3b8;background:#1e293b}
.tab.active{background:#6366f1;color:#fff;border-color:#6366f1}
.tab-panel{display:none}.tab-panel.active{display:block}
.panel-header{display:flex;align-items:center;justify-content:space-between;margin-bottom:14px}
.panel-header h3{font-size:14px;font-weight:600;color:#e2e8f0}
.staff-table{width:100%;border-collapse:collapse}
.staff-table th{text-align:left;font-size:11px;font-weight:600;text-transform:uppercase;
  letter-spacing:.08em;color:#64748b;padding:0 10px 10px 0}
.staff-table td{padding:4px 8px 4px 0;vertical-align:middle}
.cell-input{background:#0f1117;border:1px solid #2d3748;border-radius:6px;
  color:#e2e8f0;padding:7px 10px;font-size:13px;width:100%;outline:none}
.cell-input:focus{border-color:#6366f1}
.cell-input.rate{max-width:110px}
.btn-remove{background:none;border:none;color:#64748b;cursor:pointer;font-size:18px;padding:4px 8px;border-radius:4px}
.btn-remove:hover{color:#ef4444;background:#1f1515}
.btn-add{background:none;border:1px dashed #334155;color:#64748b;border-radius:7px;
  padding:6px 14px;font-size:13px;cursor:pointer}
.btn-add:hover{border-color:#6366f1;color:#6366f1}
.save-row{display:flex;align-items:center;gap:14px;margin-top:18px;
  padding-top:14px;border-top:1px solid #1e293b}
.btn-save{background:#6366f1;color:#fff;border:none;border-radius:7px;
  padding:8px 20px;font-size:13px;font-weight:500;cursor:pointer}
.btn-save:hover{background:#4f46e5}
.sheet-wrap{overflow-x:auto}
.live-table{border-collapse:collapse;font-size:12px;min-width:800px;width:100%}
.live-table th{background:#131720;color:#64748b;font-size:10px;font-weight:600;
  text-transform:uppercase;letter-spacing:.06em;padding:8px 10px;
  border-bottom:1px solid #2d3748;white-space:nowrap;text-align:center}
.live-table th.left{text-align:left}
.live-table td{padding:7px 10px;border-bottom:1px solid #1e293b;color:#cbd5e1;text-align:center}
.live-table td.name{text-align:left;font-weight:500;color:#e2e8f0}
.live-table td.rate{color:#64748b;font-size:11px}
.live-table td.hrs{color:#a5b4fc}
.live-table td.pay{color:#22c55e}
.live-table td.zero{color:#2d3748}
.live-table tr:hover td{background:#161b27}
.live-table tr.income-row td{background:#131b13;color:#4ade80;font-weight:600}
.live-table tr.total-row td{background:#131720;font-weight:600;color:#e2e8f0;border-top:2px solid #2d3748}
.hint{font-size:12px;color:#475569;margin-bottom:16px}
.file-input{background:#0f1117;border:1px solid #2d3748;border-radius:7px;
  color:#94a3b8;padding:8px 12px;font-size:13px;cursor:pointer}
.badge{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}
.badge-active{background:#134e26;color:#4ade80}
.badge-inactive{background:#1e1e1e;color:#475569}
.toast{position:fixed;bottom:20px;right:20px;background:#22c55e;color:#fff;
  padding:11px 18px;border-radius:8px;font-size:13px;font-weight:500;
  opacity:0;transition:opacity .3s;pointer-events:none;z-index:999}
.toast.show{opacity:1}.toast.error{background:#ef4444}
'''

TOAST_JS = '''
<div class="toast" id="toast"></div>
<script>
function toast(msg,isError){const t=document.getElementById('toast');t.textContent=msg;
  t.className='toast show'+(isError?' error':'');setTimeout(()=>t.className='toast',3000)}
function showTab(loc){document.querySelectorAll('.tab-panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById('panel-'+loc).classList.add('active');
  document.getElementById('tab-'+loc).classList.add('active')}
</script>'''


def topbar(active='sheet'):
    return f'''
<div class="topbar">
  <div class="topbar-left">
    <span class="logo">MB Ireland</span>
    <h1>Wages Admin</h1>
    <nav class="nav">
      <a href="/admin" class="{"active" if active=="sheet" else ""}">Live Sheet</a>
      <a href="/admin/manage" class="{"active" if active=="manage" else ""}">Staff &amp; Rates</a>
      <a href="/admin/settings" class="{"active" if active=="settings" else ""}">Settings</a>
    </nav>
  </div>
  <a href="/admin/logout" class="logout">Sign out</a>
</div>'''


# ── Routes ────────────────────────────────────────────────────────────────────

@admin_bp.route('/admin/login', methods=['GET', 'POST'])
def login():
    error = ''
    if request.method == 'POST':
        if request.form.get('password', '') == ADMIN_PASSWORD:
            resp = make_response(redirect('/admin'))
            resp.set_cookie('admin_token', ADMIN_PASSWORD, max_age=86400*7, httponly=True)
            return resp
        error = 'Incorrect password'
    return f'''<!DOCTYPE html><html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>MB Ireland — Admin</title>
<style>{STYLES}
.login-wrap{{min-height:100vh;display:flex;align-items:center;justify-content:center}}
.login-card{{background:#1a1f2e;border:1px solid #2d3748;border-radius:12px;padding:40px;width:100%;max-width:360px}}
.login-card h1{{font-size:20px;font-weight:600;color:#f1f5f9;margin-bottom:6px}}
.login-card p{{font-size:13px;color:#64748b;margin-bottom:24px}}
label{{display:block;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.08em;color:#94a3b8;margin-bottom:8px}}
input[type=password]{{width:100%;padding:10px 14px;background:#0f1117;border:1px solid #2d3748;border-radius:8px;color:#f1f5f9;font-size:14px;outline:none}}
input[type=password]:focus{{border-color:#6366f1}}
.login-btn{{width:100%;margin-top:18px;padding:11px;background:#6366f1;color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:500;cursor:pointer}}
.err{{background:#2d1f1f;border:1px solid #7f1d1d;border-radius:7px;padding:9px 13px;font-size:13px;color:#fca5a5;margin-bottom:18px}}
</style></head><body>
<div class="login-wrap"><div class="login-card">
  <div class="logo" style="margin-bottom:20px">Memory Block Ireland</div>
  <h1>Wages Admin</h1>
  <p>Sign in to manage staff and wages.</p>
  {"<div class='err'>" + error + "</div>" if error else ""}
  <form method="POST">
    <label>Password</label>
    <input type="password" name="password" autofocus placeholder="Enter password">
    <button class="login-btn" type="submit">Sign in →</button>
  </form>
</div></div></body></html>'''


@admin_bp.route('/admin/logout')
def logout():
    resp = make_response(redirect('/admin/login'))
    resp.delete_cookie('admin_token')
    return resp


@admin_bp.route('/admin')
@login_required
def admin_sheet():
    current_file = get_current_file()
    if not current_file:
        return f'''<!DOCTYPE html><html><head><style>{STYLES}</style></head>
        <body>{topbar()}<div class="main"><p style="color:#64748b;padding:40px 0">
        No wages file found. Upload a master template in Settings.</p></div></body></html>'''

    wb = load_workbook(current_file, data_only=True)

    location_tabs   = ''
    location_panels = ''

    for i, loc in enumerate(ACTIVE_SHEETS):
        if loc not in wb.sheetnames:
            continue
        active = 'active' if i == 0 else ''
        location_tabs += f'<button class="tab {active}" onclick="showTab(\'{loc}\')" id="tab-{loc}">{loc}</button>'

        ws   = wb[loc]
        data = read_sheet_data(ws)
        weeks = data['weeks']

        # Build table headers
        week_ths = ''
        for w in weeks:
            week_ths += f'<th>{w["header"]}</th>'
            if w['sun_col']:
                week_ths += f'<th>{w["sun_header"]}</th>'

        # Income row
        income_cells = ''
        for w in weeks:
            inc = sum(data['income'].get(c, 0) for c in w['wday_cols'])
            income_cells += f'<td>{"€{:,.2f}".format(inc) if inc else "—"}</td>'
            if w['sun_col']:
                income_cells += '<td>—</td>'

        # Staff rows
        staff_rows = ''
        for s in data['staff']:
            hour_cells = ''
            for w in weeks:
                wh = s['week_hours'].get(tuple(w['wday_cols']), {}).get('wday', 0)
                sh = s['week_hours'].get(tuple(w['wday_cols']), {}).get('sun', 0)
                hour_cells += f'<td class="{"hrs" if wh else "zero"}">{wh if wh else "—"}</td>'
                if w['sun_col']:
                    hour_cells += f'<td class="{"hrs" if sh else "zero"}">{sh if sh else "—"}</td>'

            rate_mon = f'€{s["rate_mon"]:.2f}' if s['rate_mon'] else '—'
            rate_sun = f'€{s["rate_sun"]:.2f}' if s['rate_sun'] else '—'
            tot_wday = f'{s["total_wday"]:.1f}h' if s['total_wday'] else '—'
            tot_sun  = f'{s["total_sun"]:.1f}h'  if s['total_sun']  else '—'
            final    = f'€{s["final"]:.2f}'       if s['final']      else '—'
            bonus    = f'€{s["bonus"]:.2f}'       if s['bonus']      else '—'
            adj      = f'€{s["adj"]:.2f}'         if s['adj']        else '—'

            staff_rows += f'''<tr>
              <td class="name">{s["name"]}</td>
              <td class="rate">{rate_mon}</td>
              <td class="rate">{rate_sun}</td>
              {hour_cells}
              <td>{tot_wday}</td>
              <td>{tot_sun}</td>
              <td>{bonus}</td><td>{adj}</td>
              <td class="pay" style="font-weight:600;color:#f1f5f9">{final}</td>
            </tr>'''

        location_panels += f'''
        <div class="tab-panel {active}" id="panel-{loc}">
          <div class="sheet-wrap">
          <table class="live-table">
            <thead>
              <tr>
                <th class="left">Name</th>
                <th>M-S Rate</th><th>Sun Rate</th>
                {week_ths}
                <th>Total<br>Wday</th><th>Total<br>Sun</th>
                <th>Bonus</th><th>Adj</th><th>Final</th>
              </tr>
            </thead>
            <tbody>
              <tr class="income-row">
                <td class="name" colspan="3">Income</td>
                {income_cells}
                <td colspan="5">—</td>
              </tr>
              {staff_rows}
            </tbody>
          </table>
          </div>
        </div>'''

    file_name = current_file.name

    return f'''<!DOCTYPE html><html><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>MB Ireland — Live Sheet</title>
<style>{STYLES}</style></head><body>
{topbar('sheet')}
<div class="main">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px">
    <div>
      <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.08em;margin-bottom:4px">Current file</div>
      <div style="font-size:14px;color:#a5b4fc;font-weight:500">{file_name}</div>
    </div>
    <div style="display:flex;gap:10px">
      <button class="btn btn-secondary" onclick="triggerRun()">▶ Run wages now</button>
      <button class="btn" style="background:#2d2000;border:1px solid #7a5000;color:#fbbf24" onclick="triggerPreliminary()">⚠ Preliminary run</button>
      <a href="/download?key={ADMIN_PASSWORD}" class="btn btn-secondary">↓ Download</a>
    </div>
  </div>

  <div style="margin-bottom:12px;display:flex;gap:8px;align-items:center">
    <div class="tabs" style="margin:0">{location_tabs}</div>
    <span style="font-size:11px;color:#475569">Active locations only</span>
  </div>
  {location_panels}
</div>
{TOAST_JS}
<script>
async function triggerRun(){{
  toast('Running wages...', false);
  await fetch('/trigger?key={ADMIN_PASSWORD}', {{method:'POST'}});
  toast('Done — check email', false);
  setTimeout(()=>location.reload(), 3000);
}}
</script>
</body></html>'''


@admin_bp.route('/admin/manage')
@login_required
def admin_manage():
    staff_data = get_staff_from_template()
    location_tabs   = ''
    location_panels = ''

    for i, loc in enumerate(ACTIVE_SHEETS):
        staff  = staff_data.get(loc, [])
        active = 'active' if i == 0 else ''
        location_tabs += f'<button class="tab {active}" onclick="showTab(\'{loc}\')" id="tab-{loc}">{loc}</button>'

        rows = ''
        for s in staff:
            rows += f'''<tr>
              <td><input class="cell-input" name="name" value="{s["name"]}"></td>
              <td><input class="cell-input rate" name="rate_mon" value="{s["rate_mon"]}"></td>
              <td><input class="cell-input rate" name="rate_sun" value="{s["rate_sun"]}"></td>
              <td><button class="btn-remove" onclick="removeRow(this)">×</button></td>
            </tr>'''

        location_panels += f'''
        <div class="tab-panel {active}" id="panel-{loc}">
          <div class="panel-header">
            <h3>{loc} <span class="badge badge-active">Active</span></h3>
            <button class="btn-add" onclick="addRow('{loc}')">+ Add staff member</button>
          </div>
          <table class="staff-table" id="table-{loc}">
            <thead><tr>
              <th>Name</th><th>Mon–Sat Rate (€/h)</th><th>Sunday Rate (€/h)</th><th></th>
            </tr></thead>
            <tbody>{rows}</tbody>
          </table>
          <div class="save-row">
            <button class="btn-save" onclick="saveLocation('{loc}')">Save {loc}</button>
            <span id="status-{loc}" style="font-size:13px;color:#22c55e"></span>
          </div>
        </div>'''

    return f'''<!DOCTYPE html><html><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>MB Ireland — Staff & Rates</title>
<style>{STYLES}</style></head><body>
{topbar('manage')}
<div class="main">
  <div class="section">
    <div class="section-title">Staff &amp; Rates — Active Locations</div>
    <p class="hint">
      Changes update the master template and take effect from next month, or immediately after regenerating.
      Default rates for new staff: €14.15/h Mon-Sat, €17.98/h Sunday.
    </p>
    <div style="margin-bottom:16px">
      <span class="badge badge-inactive" style="margin-right:8px">Liffey Valley — Sold</span>
      <span class="badge badge-inactive">Whitewater — Seasonal / Inactive</span>
    </div>
    <div class="tabs">{location_tabs}</div>
    {location_panels}
  </div>
</div>
{TOAST_JS}
<script>
function addRow(loc){{
  const tbody=document.querySelector('#table-'+loc+' tbody');
  const tr=document.createElement('tr');
  tr.innerHTML=`<td><input class="cell-input" name="name" placeholder="Full name"></td>
    <td><input class="cell-input rate" name="rate_mon" value="14.15"></td>
    <td><input class="cell-input rate" name="rate_sun" value="17.98"></td>
    <td><button class="btn-remove" onclick="removeRow(this)">×</button></td>`;
  tbody.appendChild(tr);
}}
function removeRow(btn){{btn.closest('tr').remove()}}
async function saveLocation(loc){{
  const rows=document.querySelectorAll('#table-'+loc+' tbody tr');
  const staff=[];
  rows.forEach(row=>{{
    const name=row.querySelector('[name=name]').value.trim();
    if(name) staff.push({{name,rate_mon:row.querySelector('[name=rate_mon]').value,rate_sun:row.querySelector('[name=rate_sun]').value}});
  }});
  const r=await fetch('/admin/save-staff',{{method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{location:loc,staff}})}});
  const d=await r.json();
  d.ok?toast('Saved '+loc,false):toast('Error: '+d.error,true);
}}
</script></body></html>'''


@admin_bp.route('/admin/settings')
@login_required
def admin_settings():
    files = sorted(DATA_DIR.glob('MB_Ireland_*.xlsx'), reverse=True)
    files_html = ''.join(
        f'<div style="font-size:13px;color:#6366f1;margin-top:6px">{f.name}</div>'
        for f in files[:6]
    )

    return f'''<!DOCTYPE html><html><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>MB Ireland — Settings</title>
<style>{STYLES}</style></head><body>
{topbar('settings')}
<div class="main">
  <div class="cards">
    <div class="card">
      <div class="card-label">Manual Run</div>
      <div class="card-action">Trigger wages for the current week right now.</div>
      <button class="btn btn-primary" onclick="triggerRun()">Run wages now</button>
    </div>
    <div class="card">
      <div class="card-label">Regenerate Month</div>
      <div class="card-action">Wipe this month's file and start fresh from the master template.</div>
      <button class="btn btn-danger" onclick="confirmRegenerate()">Regenerate</button>
    </div>
    <div class="card">
      <div class="card-label">Download</div>
      <div class="card-action">Download the current month's wages file.</div>
      <a href="/download?key={ADMIN_PASSWORD}" class="btn btn-secondary">Download Excel</a>
    </div>
  </div>

  <div class="section">
    <div class="section-title">Upload Master Template</div>
    <p class="hint">
      Replace the master template with an updated Excel file.
      Staff names, rates, and sheet formulas will be carried into all future months.
    </p>
    <div style="display:flex;gap:12px;align-items:center;flex-wrap:wrap">
      <input type="file" id="template-file" class="file-input" accept=".xlsx">
      <button class="btn btn-secondary" onclick="uploadTemplate()">Upload</button>
    </div>
    <div style="margin-top:16px">
      <div style="font-size:11px;color:#475569;text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px">Monthly files on server</div>
      {files_html}
    </div>
  </div>

  <div class="section">
    <div class="section-title">Location Status</div>
    <table style="width:100%;border-collapse:collapse;font-size:13px">
      <thead><tr style="color:#64748b;font-size:11px;text-transform:uppercase">
        <th style="text-align:left;padding:8px 0">Location</th>
        <th style="text-align:left;padding:8px 0">Status</th>
        <th style="text-align:left;padding:8px 0">Notes</th>
      </tr></thead>
      <tbody>
        <tr><td style="padding:8px 0;color:#e2e8f0">Blanchardstown</td><td><span class="badge badge-active">Active</span></td><td style="color:#64748b;font-size:12px"></td></tr>
        <tr><td style="padding:8px 0;color:#e2e8f0">Cork</td><td><span class="badge badge-active">Active</span></td><td style="color:#64748b;font-size:12px">Two Square locations — using Mahon Point</td></tr>
        <tr><td style="padding:8px 0;color:#e2e8f0">Nutgrove</td><td><span class="badge badge-active">Active</span></td><td style="color:#64748b;font-size:12px"></td></tr>
        <tr><td style="padding:8px 0;color:#e2e8f0">Liffey Valley</td><td><span class="badge badge-inactive">Sold</span></td><td style="color:#64748b;font-size:12px">Sheet zeroed out automatically each run</td></tr>
        <tr><td style="padding:8px 0;color:#e2e8f0">Whitewater</td><td><span class="badge badge-inactive">Inactive</span></td><td style="color:#64748b;font-size:12px">Seasonal — sheet zeroed out each run</td></tr>
      </tbody>
    </table>
  </div>
</div>
{TOAST_JS}
<script>
async function triggerRun(){{
  toast('Running wages...', false);
  await fetch('/trigger?key={ADMIN_PASSWORD}', {{method:'POST'}});
  toast('Done — check email', false);
}}
async function confirmRegenerate(){{
  if(!confirm('Wipe all hours for the current month and start fresh?')) return;
  await fetch('/regenerate?key={ADMIN_PASSWORD}');
  toast('Month regenerated', false);
}}
async function uploadTemplate(){{
  const file=document.getElementById('template-file').files[0];
  if(!file){{toast('Select a file first', true);return}}
  const fd=new FormData();fd.append('file', file);
  const r=await fetch('/upload-template?key={ADMIN_PASSWORD}', {{method:'POST',body:fd}});
  const d=await r.json();
  d.status==='uploaded'?toast('Template uploaded ✓', false):toast('Upload failed', true);
}}
</script></body></html>'''


@admin_bp.route('/admin/save-staff', methods=['POST'])
@login_required
def save_staff():
    data  = request.get_json()
    loc   = data.get('location')
    staff = data.get('staff', [])
    if not loc or loc not in ACTIVE_SHEETS:
        return jsonify({'ok': False, 'error': 'Invalid location'})
    ok, msg = save_staff_to_template(loc, staff)
    return jsonify({'ok': ok, 'message': msg})
